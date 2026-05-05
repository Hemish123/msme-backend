"""
Excel template download views for bulk upload.
Generates styled .xlsx templates for Customer and Inventory bulk imports.
"""
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


class CustomerTemplateView(APIView):
    """Download customer bulk upload Excel template."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Customer Template'

        # Styles
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
        sample_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        instruction_font = Font(name='Calibri', italic=True, color='6B7280', size=10)
        instruction_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Headers (Row 1)
        headers = [
            'Customer Type (Business/Individual)',
            'Salutation (Mr./Mrs./Ms./Dr./Prof.)',
            'First Name *',
            'Last Name',
            'Company Name',
            'Display Name *',
            'Email Address *',
            'Work Phone',
            'Mobile',
            'GST Treatment *',
            'GSTIN (GST Number)',
            'PAN Number',
            'Place of Supply *',
            'Tax Preference (Taxable/Tax Exempt)',
            'Payment Terms (Due on Receipt/Net 15/Net 30/Net 45/Net 60/Net 90)',
            'Billing Street 1',
            'Billing Street 2',
            'Billing City',
            'Billing State',
            'Billing ZIP',
            'Shipping Street 1',
            'Shipping Street 2',
            'Shipping City',
            'Shipping State',
            'Shipping ZIP',
            'Remarks',
            'Invoice Number',
            'Invoice Date (YYYY-MM-DD)',
            'Due Date (YYYY-MM-DD)',
            'Invoice Amount (INR)',
            'Paid Amount (INR)',
            'Paid Date (YYYY-MM-DD)',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Sample data (Row 2)
        sample = [
            'Business', 'Mr.', 'Raj', 'Kumar', 'Raj Traders Pvt Ltd',
            'Raj Kumar', 'raj@rajtraders.com', '+91-9876543210', '+91-9876543211',
            'Registered Business - Regular', '27AABCT1332L1ZL', 'ABCDE1234F',
            'Gujarat', 'Taxable', 'Net 30',
            '123 MG Road', 'Navrangpura', 'Ahmedabad', 'Gujarat', '380009',
            '123 MG Road', 'Navrangpura', 'Ahmedabad', 'Gujarat', '380009',
            'Regular customer',
            'INV-001', '2024-03-01', '2024-03-31', '50000.00', '50000.00', '2024-03-15',
        ]

        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = sample_fill
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)

        # Instructions row (Row 3)
        instructions = [
            '* Required fields', 'Do not modify column headers',
            'Delete this row before uploading', '', '', '', '', '', '',
            '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
            'Optional invoice data', '', '', '', '', '',
        ]

        for col_idx, val in enumerate(instructions, 1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = instruction_font
            cell.fill = instruction_fill
            cell.border = thin_border

        # Set column widths
        col_widths = [30, 30, 15, 15, 25, 20, 25, 18, 18,
                      35, 20, 15, 20, 30, 50,
                      20, 20, 15, 15, 10,
                      20, 20, 15, 15, 10, 20,
                      20, 25, 25, 25, 25, 25]
        for i, w in enumerate(col_widths, 1):
            # Convert index to column letter (A, B, ..., Z, AA, AB, ...)
            col_letter = chr(64 + i) if i <= 26 else chr(64 + ((i - 1) // 26)) + chr(64 + ((i - 1) % 26) + 1)
            ws.column_dimensions[col_letter].width = w

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Write to bytes
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customer_bulk_upload_template.xlsx"'
        return response


class InventoryTemplateView(APIView):
    """Download inventory bulk upload Excel template."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Template'

        # Styles
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
        sample_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        instruction_font = Font(name='Calibri', italic=True, color='6B7280', size=10)
        instruction_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        headers = [
            'Product Name *',
            'HSN Code *',
            'Unit (Nos/Kg/Ltr/Mtr/Box/Pcs/Set/Pair/Dozen/Other) *',
            'Unit Price (INR) *',
            'Tax % (0/5/12/18/28) *',
            'Stock Quantity *',
            'Description',
            'Customer Email (optional)',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        sample = [
            'Steel Pipe 2 inch', '7304', 'Mtr', '1250.00', '18', '100',
            'Galvanized steel pipe, 2 inch diameter', 'raj@rajtraders.com',
        ]

        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = sample_fill
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)

        instructions = [
            '* Required fields', 'Do not modify column headers',
            'Delete this row before uploading', '', '', '', '', '',
        ]

        for col_idx, val in enumerate(instructions, 1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = instruction_font
            cell.fill = instruction_fill
            cell.border = thin_border

        # Set column widths
        for i, w in enumerate([25, 15, 45, 18, 22, 18, 30, 25], 1):
            ws.column_dimensions[chr(64 + i)].width = w

        ws.freeze_panes = 'A2'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_bulk_upload_template.xlsx"'
        return response
